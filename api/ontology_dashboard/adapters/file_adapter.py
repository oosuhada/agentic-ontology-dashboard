from __future__ import annotations

import csv
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

from .models import DatasetManifest, IngestionResult, QualityRule, QuarantinedRecord
from .registry import AdapterRegistry, default_adapter_registry
from .repository import AdapterRepository


class FileAccessPolicy:
    """Restrict ingestion to explicitly configured local roots."""

    def __init__(self, allowed_roots: Iterable[str | Path]) -> None:
        roots = [Path(root).expanduser().resolve() for root in allowed_roots]
        if not roots:
            raise ValueError("at least one dataset root must be configured")
        self.allowed_roots = tuple(roots)

    def validate(self, path: Path) -> Path:
        resolved = path.expanduser().resolve(strict=True)
        if not resolved.is_file():
            raise ValueError(f"dataset source is not a file: {resolved}")
        if not any(resolved.is_relative_to(root) for root in self.allowed_roots):
            raise ValueError("dataset source is outside the configured ingestion roots")
        return resolved


class FileAdapter:
    def __init__(
        self,
        database_path: str | Path,
        *,
        allowed_roots: Iterable[str | Path],
        registry: AdapterRegistry | None = None,
        repository: AdapterRepository | None = None,
    ) -> None:
        self.repository = repository or AdapterRepository(database_path)
        self.registry = registry or default_adapter_registry()
        self.policy = FileAccessPolicy(allowed_roots)

    @staticmethod
    def _source_path(uri: str) -> Path:
        parsed = urlparse(uri)
        if parsed.scheme in {"", "file"}:
            value = unquote(parsed.path) if parsed.scheme == "file" else uri
            return Path(value)
        raise ValueError("File Adapter only supports local file paths and file:// URIs")

    @staticmethod
    def _checksum(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _resolve_headers(
        headers: list[str],
        manifest: DatasetManifest,
    ) -> tuple[dict[str, str], list[str]]:
        available = {header.strip(): header for header in headers if header is not None}
        aliases = manifest.schema_.field_aliases
        resolved: dict[str, str] = {}
        canonical_fields = set(aliases) | set(manifest.schema_.required_fields)
        for canonical in sorted(canonical_fields):
            candidates = [canonical, *aliases.get(canonical, [])]
            matched = next((available[item] for item in candidates if item in available), None)
            if matched is not None:
                resolved[canonical] = matched
        missing = [
            canonical
            for canonical in manifest.schema_.required_fields
            if canonical not in resolved
        ]
        return resolved, missing

    @staticmethod
    def _apply_aliases(record: dict[str, str], mapping: dict[str, str]) -> dict[str, str]:
        normalized = dict(record)
        for canonical, source_field in mapping.items():
            normalized[canonical] = record.get(source_field, "")
        return normalized

    @staticmethod
    def _quality_error(
        rule: QualityRule,
        value: str | None,
    ) -> str | None:
        text = "" if value is None else str(value).strip()
        if rule.rule == "required":
            return None if text else "value is required"
        if not text:
            return None
        try:
            if rule.rule == "number":
                float(text)
            elif rule.rule == "integer":
                parsed = float(text)
                if not parsed.is_integer():
                    return "value is not an integer"
            elif rule.rule == "datetime":
                datetime.fromisoformat(text.replace("Z", "+00:00"))
            elif rule.rule == "enum":
                allowed = set(rule.value if isinstance(rule.value, list) else [])
                if text not in allowed:
                    return f"value is not in {sorted(allowed)}"
            elif rule.rule == "min" and float(text) < float(rule.value):
                return f"value is below {rule.value}"
            elif rule.rule == "max" and float(text) > float(rule.value):
                return f"value is above {rule.value}"
        except (TypeError, ValueError) as exc:
            return str(exc)
        return None

    def _normalize_rows(
        self,
        *,
        headers: list[str],
        rows: Iterable[tuple[int, dict[str, Any]]],
        manifest: DatasetManifest,
    ) -> tuple[int, list[dict[str, Any]], list[QuarantinedRecord]]:
        adapter = self.registry.get(manifest.adapter_code)
        alias_mapping, missing = self._resolve_headers(headers, manifest)
        adapter_missing = sorted(adapter.required_fields(manifest) - set(alias_mapping))
        missing = sorted(set(missing) | set(adapter_missing))
        if missing:
            raise ValueError(f"dataset is missing required fields: {', '.join(missing)}")
        accepted: list[dict[str, Any]] = []
        quarantined: list[QuarantinedRecord] = []
        source_count = 0
        for row_number, raw_record in rows:
            source_count += 1
            record = self._apply_aliases(
                {
                    key: "" if value is None else str(value)
                    for key, value in raw_record.items()
                    if key is not None
                },
                alias_mapping,
            )
            quality_issue = next(
                (
                    (rule, message)
                    for rule in manifest.quality_rules
                    if (message := self._quality_error(rule, record.get(rule.field)))
                ),
                None,
            )
            if quality_issue is not None:
                rule, message = quality_issue
                quarantined.append(
                    QuarantinedRecord(
                        source_row_number=row_number,
                        error_code=rule.code,
                        error_message=message,
                        record=record,
                    )
                )
                continue
            normalized = adapter.normalize_record(
                record,
                manifest=manifest,
                row_number=row_number,
            )
            if isinstance(normalized, QuarantinedRecord):
                quarantined.append(normalized)
            else:
                accepted.append(normalized)
        return source_count, accepted, quarantined

    def ingest(self, manifest: DatasetManifest) -> IngestionResult:
        path = self.policy.validate(self._source_path(manifest.source.uri))
        checksum = self._checksum(path)
        if checksum.lower() != manifest.source.checksum_sha256.lower():
            raise ValueError("dataset checksum does not match the manifest")
        if manifest.schema_.format not in {"csv", "xlsx"}:
            raise ValueError("the current File Adapter runtime supports CSV and XLSX manifests")

        adapter = self.registry.get(manifest.adapter_code)
        self.repository.save_manifest(manifest)
        run_id = self.repository.start_run(manifest)
        accepted: list[dict[str, Any]] = []
        quarantined: list[QuarantinedRecord] = []
        source_count = 0
        try:
            if manifest.schema_.format == "csv":
                with path.open(
                    "r",
                    encoding=manifest.source.encoding or "utf-8",
                    newline="",
                ) as handle:
                    reader = csv.DictReader(
                        handle,
                        delimiter=manifest.schema_.delimiter or ",",
                    )
                    if reader.fieldnames is None:
                        raise ValueError("CSV header is required")
                    source_count, accepted, quarantined = self._normalize_rows(
                        headers=[str(item) for item in reader.fieldnames],
                        rows=(
                            (row_number, raw_record)
                            for row_number, raw_record in enumerate(reader, start=2)
                        ),
                        manifest=manifest,
                    )
            else:
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    if manifest.schema_.sheet:
                        if manifest.schema_.sheet not in workbook.sheetnames:
                            raise ValueError(
                                f"XLSX sheet not found: {manifest.schema_.sheet}"
                            )
                        worksheet = workbook[manifest.schema_.sheet]
                    else:
                        worksheet = workbook.active
                    values = worksheet.iter_rows(values_only=True)
                    header_row = next(values, None)
                    if header_row is None:
                        raise ValueError("XLSX header is required")
                    headers = ["" if value is None else str(value).strip() for value in header_row]
                    if not any(headers):
                        raise ValueError("XLSX header is required")
                    source_count, accepted, quarantined = self._normalize_rows(
                        headers=headers,
                        rows=(
                            (
                                row_number,
                                {
                                    header: value
                                    for header, value in zip(headers, row, strict=False)
                                    if header
                                },
                            )
                            for row_number, row in enumerate(values, start=2)
                        ),
                        manifest=manifest,
                    )
                finally:
                    workbook.close()
            metrics = adapter.derive_metrics(accepted, manifest=manifest)
            self.repository.complete_run(
                run_id=run_id,
                manifest=manifest,
                source_count=source_count,
                accepted_count=len(accepted),
                quarantined=quarantined,
            )
        except Exception as exc:
            self.repository.complete_run(
                run_id=run_id,
                manifest=manifest,
                source_count=source_count,
                accepted_count=len(accepted),
                quarantined=quarantined,
                error_message=str(exc),
            )
            raise

        return IngestionResult(
            ingestion_run_id=run_id,
            manifest_id=manifest.manifest_id,
            organization_id=manifest.organization_id,
            project_id=manifest.project_id,
            workspace_id=manifest.workspace_id,
            adapter_code=manifest.adapter_code,
            status="completed_with_quarantine" if quarantined else "completed",
            source_record_count=source_count,
            accepted_record_count=len(accepted),
            quarantined_record_count=len(quarantined),
            accepted_records=accepted,
            quarantined_records=quarantined,
            metrics=metrics,
        )


__all__ = ["FileAccessPolicy", "FileAdapter", "IngestionResult"]
