from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from ontology_dashboard.migrations import migrate
from ontology_dashboard.modeling.intake import DatasetIntakeProfiler, REDACTED
from ontology_dashboard.modeling.models import (
    ManifestDraftDecisionRequest,
    ManifestDraftUpdateRequest,
    ManifestFieldSuggestion,
)
from ontology_dashboard.modeling.repository import ModelingRepository
from ontology_dashboard.modeling.service import ModelingService


class InvalidProvider:
    def generate_json(self, system_prompt: str, payload: dict) -> dict:
        return {
            "structure_type": "wide_pivot",
            "selected_fields": ["invented_field"],
            "rationale": "invalid on purpose",
        }


def service(tmp_path: Path, *, provider=None) -> ModelingService:
    database = tmp_path / "intake.db"
    migrate(str(database))
    profiler = DatasetIntakeProfiler([tmp_path / "sources"], provider=provider)
    return ModelingService(
        ModelingRepository(database),
        intake_profiler=profiler,
    )


def write_csv(path: Path, rows: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def profile_csv(service_: ModelingService, path: Path, key: str = "profile"):
    return service_.profile_source(
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
        source_path=str(path),
        sheet=None,
        use_llm=True,
        idempotency_key=key,
        actor_id="user-fde",
    )


def test_csv_delimiter_profile_redaction_and_manifest_approval(tmp_path: Path) -> None:
    path = tmp_path / "sources" / "telemetry.csv"
    write_csv(
        path,
        [
            "machine_id;datetime;voltage;api_token",
            "M-001;2026-01-01T00:00:00;221.5;secret-a",
            "M-001;2026-01-01T00:10:00;222.1;secret-b",
        ],
    )
    modeling = service(tmp_path)
    profile = profile_csv(modeling, path)
    assert profile.status == "ready_for_review"
    assert profile.structure_type == "tabular_column_as_attribute"
    assert profile.row_count == 2
    assert profile.preview_rows[0]["api_token"] == REDACTED
    sensitive = next(item for item in profile.field_profiles if item.name == "api_token")
    assert sensitive.potential_sensitive is True
    assert sensitive.summary["redacted"] is True
    assert "secret-a" not in json.dumps(profile.model_dump(mode="json"))
    voltage = next(item for item in profile.field_profiles if item.name == "voltage")
    assert voltage.inferred_datatype == "number"
    assert voltage.summary["min"] == 221.5

    draft = modeling.create_manifest_draft(
        profile_id=profile.profile_id,
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
        idempotency_key="draft-a",
        actor_id="user-fde",
    )
    assert draft.status == "draft"
    with pytest.raises(ValueError, match="approved Manifest Draft"):
        modeling.approved_manifest_payload(
            draft.draft_id,
            organization_id="org-a",
            project_id="project-a",
            workspace_id="workspace-a",
        )
    approved = modeling.decide_manifest_draft(
        draft.draft_id,
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
        request=ManifestDraftDecisionRequest(
            project_id="project-a",
            workspace_id="workspace-a",
            expected_revision=1,
            decision="approve",
            rationale="identifier and timestamp reviewed",
        ),
        actor_id="user-fde",
    )
    assert approved.status == "approved"
    manifest = modeling.approved_manifest_payload(
        draft.draft_id,
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
    )
    assert manifest["source"]["checksum_sha256"] == profile.source_checksum_sha256
    assert manifest["approval"]["approved_by"] == "user-fde"


def test_same_preview_different_tail_has_different_full_checksum_cache_and_identity(tmp_path: Path) -> None:
    prefix = ["machine_id,datetime,value"] + [
        f"M-001,2026-01-01T00:{index:02d}:00,{index}" for index in range(40)
    ]
    first = tmp_path / "sources" / "first.csv"
    second = tmp_path / "sources" / "second.csv"
    write_csv(first, [*prefix, "M-001,2026-01-02T00:00:00,100"])
    write_csv(second, [*prefix, "M-001,2026-01-02T00:00:00,999"])
    modeling = service(tmp_path)
    left = profile_csv(modeling, first, "first")
    right = profile_csv(modeling, second, "second")
    assert left.preview_rows == right.preview_rows
    assert left.source_checksum_sha256 != right.source_checksum_sha256
    assert left.cache_key != right.cache_key
    assert left.profile_id != right.profile_id


def test_tsv_and_invalid_llm_fall_back_to_deterministic_structure(tmp_path: Path) -> None:
    path = tmp_path / "sources" / "telemetry.tsv"
    write_csv(
        path,
        [
            "equipment_id\ttimestamp\tpressure",
            "EQ-1\t2026-01-01T00:00:00\t2.1",
            "EQ-1\t2026-01-01T00:10:00\t2.2",
        ],
    )
    profile = profile_csv(service(tmp_path, provider=InvalidProvider()), path)
    assert profile.structure_type == "tabular_column_as_attribute"
    assert [item.name for item in profile.field_profiles] == [
        "equipment_id",
        "timestamp",
        "pressure",
    ]


def test_xlsx_multi_header_and_sheet_selection(tmp_path: Path) -> None:
    source = tmp_path / "sources" / "multi.xlsx"
    source.parent.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Telemetry"
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "identity"
    worksheet["C1"] = "sensor"
    worksheet.append(["machine_id", "timestamp", "voltage"])
    worksheet.append(["M-1", "2026-01-01T00:00:00", 220.0])
    workbook.create_sheet("Other").append(["not", "selected"])
    workbook.save(source)
    modeling = service(tmp_path)
    profile = modeling.profile_source(
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
        source_path=str(source),
        sheet="Telemetry",
        use_llm=False,
        idempotency_key="xlsx",
        actor_id="user-fde",
    )
    assert profile.structure_type == "multi_header"
    assert profile.row_count == 1
    assert profile.field_profiles[0].name.startswith("identity")
    with pytest.raises(ValueError, match="unknown XLSX sheet"):
        modeling.intake_profiler.profile(
            organization_id="org-a",
            project_id="project-a",
            workspace_id="workspace-a",
            source_path=str(source),
            sheet="Missing",
            use_llm=False,
            idempotency_key="missing-sheet",
        )


def test_allowed_root_and_symlink_traversal_are_rejected(tmp_path: Path) -> None:
    source_root = tmp_path / "sources"
    source_root.mkdir()
    outside = tmp_path / "outside.csv"
    write_csv(outside, ["id,value", "1,2"])
    profiler = DatasetIntakeProfiler([source_root])
    with pytest.raises(ValueError, match="outside configured allowed roots"):
        profiler.resolve_source(str(outside))
    link = source_root / "escape.csv"
    link.symlink_to(outside)
    with pytest.raises(ValueError, match="outside configured allowed roots"):
        profiler.resolve_source(str(link))


def test_missing_prerequisites_block_approval_until_explicit_review(tmp_path: Path) -> None:
    path = tmp_path / "sources" / "measures.csv"
    write_csv(path, ["value,pressure", "1,2", "2,3"])
    modeling = service(tmp_path)
    profile = profile_csv(modeling, path)
    draft = modeling.create_manifest_draft(
        profile_id=profile.profile_id,
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
        idempotency_key="draft-missing",
        actor_id="user-fde",
    )
    assert set(draft.missing_prerequisites) == {"equipment identifier", "ordered timestamp"}
    with pytest.raises(ValueError, match="unresolved prerequisites"):
        modeling.decide_manifest_draft(
            draft.draft_id,
            organization_id="org-a",
            project_id="project-a",
            workspace_id="workspace-a",
            request=ManifestDraftDecisionRequest(
                project_id="project-a",
                workspace_id="workspace-a",
                expected_revision=1,
                decision="approve",
                rationale="not enough metadata",
            ),
            actor_id="user-fde",
        )
    updated = modeling.update_manifest_draft(
        draft.draft_id,
        organization_id="org-a",
        project_id="project-a",
        workspace_id="workspace-a",
        request=ManifestDraftUpdateRequest(
            project_id="project-a",
            workspace_id="workspace-a",
            expected_revision=1,
            field_suggestions=[
                ManifestFieldSuggestion(
                    source_field="value",
                    canonical_field="equipment_id",
                    selected=True,
                    required=True,
                    essential_key=True,
                    rationale="explicit user mapping",
                    confidence=1.0,
                ),
                ManifestFieldSuggestion(
                    source_field="pressure",
                    canonical_field="observed_at",
                    selected=True,
                    required=True,
                    essential_key=True,
                    rationale="explicit user mapping",
                    confidence=1.0,
                ),
            ],
        ),
        actor_id="user-fde",
    )
    assert updated.revision == 2
    # Explicit field edits do not silently clear prerequisite warnings. The user
    # must revise the source/profile rather than forcing an unsafe approval.
    assert updated.missing_prerequisites
