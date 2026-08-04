from __future__ import annotations

import csv
import hashlib
import io
import math
import mimetypes
import re
import statistics
import uuid
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, ValidationError

from .models import (
    DatasetIntakeProfile,
    FieldProfile,
    ManifestDraft,
    ManifestFieldSuggestion,
    canonical_checksum,
)

PARSER_VERSION = "dataset-intake-v1"
MAX_PREVIEW_ROWS = 40
MAX_PREVIEW_BYTES = 256 * 1024
REDACTED = "***REDACTED***"

SENSITIVE_NAME = re.compile(
    r"(?:password|passwd|secret|token|api[_-]?key|authorization|credential|ssn|resident|email|phone)",
    re.IGNORECASE,
)
TIMESTAMP_NAME = re.compile(r"(?:time|timestamp|datetime|date|observed_at|event_at)", re.IGNORECASE)
IDENTIFIER_NAME = re.compile(
    r"(?:^id$|_id$|machine|equipment|asset|device|serial|uuid|uid)", re.IGNORECASE
)
GROUP_NAME = re.compile(r"(?:machine|equipment|asset|device|site|cell|line).*id", re.IGNORECASE)
UNIT_HINTS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"temperature|temp", re.IGNORECASE), "temperature"),
    (re.compile(r"rpm|rotation|speed", re.IGNORECASE), "rpm"),
    (re.compile(r"torque", re.IGNORECASE), "N·m"),
    (re.compile(r"pressure", re.IGNORECASE), "pressure"),
    (re.compile(r"voltage|volt", re.IGNORECASE), "V"),
    (re.compile(r"vibration", re.IGNORECASE), "vibration"),
    (re.compile(r"wear", re.IGNORECASE), "minute"),
]


class StructureSuggestion(BaseModel):
    model_config = ConfigDict(extra="forbid")
    structure_type: str
    selected_fields: list[str]
    rationale: str


class IntakeLLMProvider(Protocol):
    def generate_json(self, system_prompt: str, payload: dict[str, Any]) -> dict[str, Any]: ...


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parse_datetime(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    if not any(character in text for character in ("-", "/", ":", "T")):
        return False
    try:
        datetime.fromisoformat(text.replace("Z", "+00:00"))
        return True
    except ValueError:
        return False


def _infer_type(values: list[Any]) -> str:
    clean = [value for value in values if value not in (None, "")]
    if not clean:
        return "unknown"
    bool_values = {"true", "false", "yes", "no", "0", "1"}
    if all(str(value).strip().lower() in bool_values for value in clean):
        return "boolean"
    try:
        numbers = [float(str(value).strip()) for value in clean]
        return "integer" if all(number.is_integer() for number in numbers) else "number"
    except ValueError:
        pass
    if sum(_parse_datetime(str(value)) for value in clean) / len(clean) >= 0.8:
        return "datetime"
    return "string"


def _safe_numeric_summary(values: list[Any]) -> dict[str, Any]:
    numbers: list[float] = []
    for value in values:
        if value in (None, ""):
            continue
        try:
            numbers.append(float(str(value).strip()))
        except ValueError:
            continue
    if not numbers:
        return {}
    ordered = sorted(numbers)
    midpoint = len(ordered) // 2
    return {
        "min": min(numbers),
        "max": max(numbers),
        "mean": statistics.fmean(numbers),
        "median": statistics.median(numbers),
        "q25": ordered[max(0, math.floor((len(ordered) - 1) * 0.25))],
        "q75": ordered[max(0, math.floor((len(ordered) - 1) * 0.75))],
        "sample_count": len(numbers),
        "midpoint_index": midpoint,
    }


def _unit_hint(name: str) -> str | None:
    return next((unit for pattern, unit in UNIT_HINTS if pattern.search(name)), None)


def _semantic_candidates(name: str, inferred_type: str, distinct_count: int, row_count: int) -> list[str]:
    candidates: list[str] = []
    if TIMESTAMP_NAME.search(name) or inferred_type == "datetime":
        candidates.append("timestamp")
    if IDENTIFIER_NAME.search(name):
        candidates.append("identifier")
    if GROUP_NAME.search(name):
        candidates.append("group_key")
    if inferred_type in {"integer", "number"}:
        candidates.append("measure")
    elif inferred_type == "string":
        if row_count and distinct_count / row_count > 0.8:
            candidates.append("text")
        else:
            candidates.append("dimension")
    return list(dict.fromkeys(candidates))


def _redact_row(row: dict[str, Any], sensitive_fields: set[str]) -> dict[str, Any]:
    return {
        key: REDACTED if key in sensitive_fields and value not in (None, "") else value
        for key, value in row.items()
    }


def _normalize_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: Counter[str] = Counter()
    for index, value in enumerate(values):
        base = str(value).strip() if value not in (None, "") else f"column_{index + 1}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def classify_structure(headers: list[str], rows: list[dict[str, Any]], *, multi_header: bool = False) -> str:
    if multi_header:
        return "multi_header"
    if not headers:
        return "unsupported"
    lowered = [header.lower() for header in headers]
    if len(headers) == 2 and (
        set(lowered) & {"key", "property", "attribute", "field"}
        and set(lowered) & {"value", "val"}
    ):
        return "key_value"
    date_like_headers = sum(_parse_datetime(header) for header in headers)
    if len(headers) >= 8 and date_like_headers >= max(3, len(headers) // 2):
        return "wide_pivot"
    if len(headers) >= 3 and lowered[0] in {"property", "attribute", "metric", "field"}:
        return "tabular_row_as_attribute"
    if rows:
        first_column = [row.get(headers[0]) for row in rows]
        if len(headers) >= 3 and all(isinstance(value, str) for value in first_column if value is not None):
            unique_ratio = len({str(value) for value in first_column}) / max(1, len(first_column))
            numeric_other = sum(
                _infer_type([row.get(header) for row in rows]) in {"integer", "number"}
                for header in headers[1:]
            )
            if unique_ratio >= 0.8 and numeric_other >= max(1, len(headers[1:]) // 2):
                return "tabular_column_as_attribute"
    return "tabular_column_as_attribute"


class DatasetIntakeProfiler:
    def __init__(
        self,
        allowed_roots: list[str | Path],
        *,
        provider: IntakeLLMProvider | None = None,
        parser_version: str = PARSER_VERSION,
    ) -> None:
        roots = [Path(root).expanduser().resolve() for root in allowed_roots]
        if not roots:
            raise ValueError("at least one Dataset Intake root must be configured")
        self.allowed_roots = tuple(roots)
        self.provider = provider
        self.parser_version = parser_version

    def resolve_source(self, source_path: str) -> Path:
        candidate = Path(source_path).expanduser().resolve(strict=True)
        if not candidate.is_file():
            raise ValueError("Dataset Intake source must be a file")
        if not any(candidate.is_relative_to(root) for root in self.allowed_roots):
            raise ValueError("Dataset Intake source is outside configured allowed roots")
        return candidate

    def _read_csv(self, path: Path) -> tuple[list[str], list[dict[str, Any]], int, dict[str, Any]]:
        raw = path.read_bytes()
        bounded = raw[:MAX_PREVIEW_BYTES]
        encoding = "utf-8-sig"
        try:
            text = bounded.decode(encoding)
        except UnicodeDecodeError:
            encoding = "cp949"
            try:
                text = bounded.decode(encoding)
            except UnicodeDecodeError:
                encoding = "latin-1"
                text = bounded.decode(encoding)
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
        stream = io.StringIO(text)
        reader = csv.reader(stream, delimiter=delimiter)
        try:
            headers = _normalize_headers(next(reader))
        except StopIteration:
            return [], [], 0, {"encoding": encoding, "delimiter": delimiter}
        rows: list[dict[str, Any]] = []
        for values in reader:
            if len(rows) >= MAX_PREVIEW_ROWS:
                break
            padded = list(values[: len(headers)]) + [""] * max(0, len(headers) - len(values))
            rows.append(dict(zip(headers, padded, strict=True)))
        with path.open("r", encoding=encoding, newline="", errors="replace") as handle:
            row_count = max(0, sum(1 for _ in handle) - 1)
        return headers, rows, row_count, {"encoding": encoding, "delimiter": delimiter}

    def _read_xlsx(
        self,
        path: Path,
        sheet: str | None,
    ) -> tuple[list[str], list[dict[str, Any]], int, dict[str, Any]]:
        workbook = load_workbook(path, read_only=False, data_only=True)
        if sheet is not None and sheet not in workbook.sheetnames:
            raise ValueError(f"unknown XLSX sheet: {sheet}")
        worksheet = workbook[sheet or workbook.sheetnames[0]]
        values = worksheet.iter_rows(values_only=True)
        try:
            first = list(next(values))
        except StopIteration:
            return [], [], 0, {"sheet": worksheet.title, "multi_header": False}
        multi_header = bool(worksheet.merged_cells.ranges)
        second: list[Any] | None = None
        if not multi_header:
            try:
                second = list(next(values))
            except StopIteration:
                second = None
            if second is not None:
                first_nonempty = sum(value not in (None, "") for value in first)
                second_nonempty = sum(value not in (None, "") for value in second)
                first_numeric = sum(isinstance(value, (int, float)) for value in first)
                if first_nonempty < second_nonempty and first_numeric == 0:
                    multi_header = True
        if multi_header and second is not None:
            headers = _normalize_headers(
                [
                    "__".join(str(part).strip() for part in pair if part not in (None, ""))
                    or f"column_{index + 1}"
                    for index, pair in enumerate(zip(first, second, strict=False))
                ]
            )
        else:
            headers = _normalize_headers(first)
        rows: list[dict[str, Any]] = []
        if second is not None and not multi_header:
            padded = list(second[: len(headers)]) + [None] * max(0, len(headers) - len(second))
            rows.append(dict(zip(headers, padded, strict=True)))
        for values_row in values:
            if len(rows) >= MAX_PREVIEW_ROWS:
                break
            values_list = list(values_row)
            padded = values_list[: len(headers)] + [None] * max(0, len(headers) - len(values_list))
            rows.append(dict(zip(headers, padded, strict=True)))
        row_count = max(0, worksheet.max_row - (2 if multi_header else 1))
        return headers, rows, row_count, {"sheet": worksheet.title, "multi_header": multi_header}

    def profile(
        self,
        *,
        organization_id: str,
        project_id: str,
        workspace_id: str,
        source_path: str,
        sheet: str | None,
        use_llm: bool,
        idempotency_key: str,
    ) -> DatasetIntakeProfile:
        path = self.resolve_source(source_path)
        checksum = sha256_file(path)
        cache_key = canonical_checksum(
            {"source_checksum_sha256": checksum, "parser_version": self.parser_version}
        )
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv", ".txt"}:
            headers, rows, row_count, parser_metadata = self._read_csv(path)
            media_type = mimetypes.guess_type(path.name)[0] or "text/csv"
        elif suffix in {".xlsx", ".xlsm"}:
            headers, rows, row_count, parser_metadata = self._read_xlsx(path, sheet)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            headers, rows, row_count, parser_metadata = [], [], 0, {}
            media_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        structure = classify_structure(
            headers,
            rows,
            multi_header=bool(parser_metadata.get("multi_header")),
        )
        sensitive_fields = {header for header in headers if SENSITIVE_NAME.search(header)}
        field_profiles: list[FieldProfile] = []
        for header in headers:
            values = [row.get(header) for row in rows]
            inferred_type = _infer_type(values)
            distinct_count = len({str(value) for value in values if value not in (None, "")})
            semantics = _semantic_candidates(header, inferred_type, distinct_count, len(rows))
            sensitive = header in sensitive_fields
            summary: dict[str, Any] = {
                "unit_hint": _unit_hint(header),
                "non_null_sample_count": sum(value not in (None, "") for value in values),
                "redacted": sensitive,
            }
            if not sensitive and inferred_type in {"integer", "number"}:
                summary.update(_safe_numeric_summary(values))
            elif not sensitive:
                summary["sample_patterns"] = [
                    {"length": len(str(value)), "type": type(value).__name__}
                    for value in values[:5]
                    if value not in (None, "")
                ]
            field_profiles.append(
                FieldProfile(
                    name=header,
                    inferred_datatype=inferred_type,
                    null_ratio=(
                        sum(value in (None, "") for value in values) / len(values) if values else 1.0
                    ),
                    distinct_estimate=distinct_count,
                    semantic_candidates=semantics,
                    potential_sensitive=sensitive,
                    essential_key_candidate=bool(
                        {"identifier", "timestamp", "group_key"}.intersection(semantics)
                    ),
                    summary=summary,
                )
            )
        if use_llm and self.provider is not None and headers:
            try:
                suggestion = StructureSuggestion.model_validate(
                    self.provider.generate_json(
                        "Return a registered structure type and only existing selected fields.",
                        {
                            "registered_structure_types": [
                                "tabular_column_as_attribute",
                                "tabular_row_as_attribute",
                                "wide_pivot",
                                "key_value",
                                "multi_header",
                                "unsupported",
                            ],
                            "fields": headers,
                            "deterministic_structure": structure,
                            "preview": [_redact_row(row, sensitive_fields) for row in rows[:10]],
                        },
                    )
                )
                if suggestion.structure_type not in {
                    "tabular_column_as_attribute",
                    "tabular_row_as_attribute",
                    "wide_pivot",
                    "key_value",
                    "multi_header",
                    "unsupported",
                } or not set(suggestion.selected_fields).issubset(headers):
                    raise ValueError("LLM returned values outside the registry")
                structure = suggestion.structure_type
            except (ValidationError, ValueError, TypeError):
                pass
        unsupported = structure == "unsupported" or suffix not in {".csv", ".tsv", ".txt", ".xlsx", ".xlsm"}
        profile_id = f"intake-{uuid.uuid5(uuid.NAMESPACE_URL, f'{organization_id}:{project_id}:{workspace_id}:{cache_key}')}"
        return DatasetIntakeProfile(
            organization_id=organization_id,
            project_id=project_id,
            workspace_id=workspace_id,
            profile_id=profile_id,
            source_uri=path.as_uri(),
            source_checksum_sha256=checksum,
            parser_version=self.parser_version,
            cache_key=cache_key,
            byte_size=path.stat().st_size,
            media_type=media_type,
            status="unsupported" if unsupported else "ready_for_review",
            structure_type="unsupported" if unsupported else structure,
            field_profiles=field_profiles,
            preview_rows=[_redact_row(row, sensitive_fields) for row in rows],
            row_count=row_count,
            retryable=False,
            failure_reason=("unsupported source format" if unsupported else None),
            idempotency_key=idempotency_key,
        )


def draft_from_profile(profile: DatasetIntakeProfile, *, idempotency_key: str) -> ManifestDraft:
    if profile.status != "ready_for_review":
        raise ValueError("Manifest Draft requires a ready_for_review intake profile")
    suggestions: list[ManifestFieldSuggestion] = []
    for field in profile.field_profiles:
        semantic = set(field.semantic_candidates)
        canonical: str | None = None
        if "timestamp" in semantic:
            canonical = "observed_at"
        elif "identifier" in semantic or "group_key" in semantic:
            canonical = "equipment_id"
        suggestions.append(
            ManifestFieldSuggestion(
                source_field=field.name,
                canonical_field=canonical,
                selected=True,
                required=field.essential_key_candidate,
                rationale=(
                    "protected essential key candidate"
                    if field.essential_key_candidate
                    else f"profiled datatype={field.inferred_datatype}"
                ),
                confidence=0.95 if canonical else 0.65,
                essential_key=field.essential_key_candidate,
            )
        )
    required_semantics = {
        semantic
        for field in profile.field_profiles
        for semantic in field.semantic_candidates
        if semantic in {"identifier", "timestamp"}
    }
    missing = [
        label
        for label, semantic in (("equipment identifier", "identifier"), ("ordered timestamp", "timestamp"))
        if semantic not in required_semantics
    ]
    source_path = Path(profile.source_uri.removeprefix("file://"))
    format_name = "xlsx" if source_path.suffix.lower() in {".xlsx", ".xlsm"} else "csv"
    draft_id = f"manifest-draft-{uuid.uuid5(uuid.NAMESPACE_URL, f'{profile.profile_id}:{idempotency_key}')}"
    return ManifestDraft(
        organization_id=profile.organization_id,
        project_id=profile.project_id,
        workspace_id=profile.workspace_id,
        draft_id=draft_id,
        profile_id=profile.profile_id,
        source_checksum_sha256=profile.source_checksum_sha256,
        format=format_name,
        field_suggestions=suggestions,
        quality_rules=[
            {"field": item.source_field, "rule": "required", "code": "REQUIRED"}
            for item in suggestions
            if item.required
        ],
        missing_prerequisites=missing,
        idempotency_key=idempotency_key,
    )


__all__ = [
    "DatasetIntakeProfiler",
    "IntakeLLMProvider",
    "PARSER_VERSION",
    "REDACTED",
    "classify_structure",
    "draft_from_profile",
    "sha256_file",
]
